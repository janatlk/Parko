"""
Утилиты для экспорта отчетов в различные форматы (CSV, XLSX, PDF, JSON).
"""
import csv
import io
import json
from datetime import datetime
from typing import Any, Dict, List, Optional

from django.http import HttpResponse
from django.utils.translation import gettext as _


def export_to_csv(
    data: List[Dict[str, Any]], 
    filename: str,
    delimiter: str = ','
) -> HttpResponse:
    """
    Экспорт данных в CSV формат.
    
    Args:
        data: Список словарей с данными для экспорта
        filename: Имя файла для скачивания
    
    Returns:
        HttpResponse с CSV файлом
    """
    if not data:
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    output = io.StringIO()

    # Получаем заголовки из первой строки
    fieldnames = list(data[0].keys())

    writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=delimiter)
    writer.writeheader()
    writer.writerows(data)

    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_to_xlsx(data: List[Dict[str, Any]], filename: str, sheet_name: str = 'Report') -> HttpResponse:
    """
    Экспорт данных в XLSX формат.
    
    Args:
        data: Список словарей с данными для экспорта
        filename: Имя файла для скачивания
        sheet_name: Название листа в Excel
    
    Returns:
        HttpResponse с XLSX файлом
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    # Заголовки
    fieldnames = list(data[0].keys())
    
    # Стилизация заголовков
    for col_num, fieldname in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_num, value=fieldname)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    for row_num, row_data in enumerate(data, 2):
        for col_num, fieldname in enumerate(fieldnames, 1):
            ws.cell(row=row_num, column=col_num, value=row_data.get(fieldname))
    
    # Автоматическая ширина колонок
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_to_json(
    data: List[Dict[str, Any]], 
    filename: str,
    metadata: Optional[Dict[str, Any]] = None
) -> HttpResponse:
    """
    Экспорт данных в JSON формат.

    Args:
        data: Список словарей с данными для экспорта
        filename: Имя файла для скачивания
        metadata: Дополнительные метаданные (генерация, компания и т.д.)

    Returns:
        HttpResponse с JSON файлом
    """
    export_data = {
        'generated_at': datetime.now().isoformat(),
        'data': data,
    }
    
    if metadata:
        export_data['metadata'] = metadata

    response = HttpResponse(
        json.dumps(export_data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_to_pdf(
    data: List[Dict[str, Any]], 
    filename: str,
    title: str = 'Report',
    company_name: Optional[str] = None,
    header_info: Optional[Dict[str, str]] = None
) -> HttpResponse:
    """
    Экспорт данных в PDF формат.

    Args:
        data: Список словарей с данными для экспорта
        filename: Имя файла для скачивания
        title: Заголовок отчета
        company_name: Название компании
        header_info: Дополнительная информация в шапке (период, фильтры)

    Returns:
        HttpResponse с PDF файлом
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        raise ImportError("reportlab не установлен. Установите: pip install reportlab")

    buffer = io.BytesIO()
    
    # Создаем документ с альбомной ориентацией для широких таблиц
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    elements = []
    
    # Стили
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        alignment=1,  # Center
        spaceAfter=12
    )
    
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 12
    
    # Заголовок компании
    if company_name:
        elements.append(Paragraph(f"<b>{company_name}</b>", title_style))
        elements.append(Spacer(1, 0.2*cm))
    
    # Заголовок отчета
    elements.append(Paragraph(f"<b>{title}</b>", title_style))
    elements.append(Spacer(1, 0.2*cm))
    
    # Дополнительная информация
    if header_info:
        for key, value in header_info.items():
            elements.append(Paragraph(f"<b>{key}:</b> {value}", normal_style))
        elements.append(Spacer(1, 0.2*cm))
    
    # Таблица с данными
    if data:
        # Получаем заголовки
        headers = list(data[0].keys())
        
        # Создаем данные таблицы
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, '')) for h in headers])
        
        # Создаем таблицу
        table = Table(table_data, repeatRows=1)
        
        # Стиль таблицы
        table.setStyle(TableStyle([
            # Заголовки
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Чередование цветов строк
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
            
            # Границы
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
            
            # Размер шрифта для данных
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("<i>No data available</i>", normal_style))
    
    # Итоговая строка (если есть данные с итогами)
    if data and len(data) > 1:
        # Проверяем, есть ли строка с итогами (последняя строка с 'ИТОГО' или 'TOTAL')
        last_row = data[-1]
        if any('ИТОГО' in str(v) or 'TOTAL' in str(v) for v in last_row.values()):
            elements.append(Spacer(1, 0.3*cm))
            elements.append(Paragraph("<b>Total row included in table</b>", normal_style))
    
    # Генерируем PDF
    doc.build(elements)
    
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def prepare_maintenance_costs_for_export(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Подготовка данных отчета по затратам на ТО для экспорта.
    
    Args:
        report_data: Данные отчета из get_maintenance_costs_report
    
    Returns:
        Список словарей для экспорта
    """
    rows = []
    
    # Добавляем строки по машинам
    for car_data in report_data.get('by_car', []):
        rows.append({
            'Машина (номер)': car_data.get('car__numplate', ''),
            'Стоимость запчастей (сом)': car_data.get('part_total', 0),
            'Стоимость работ (сом)': car_data.get('job_total', 0),
            'Итого (сом)': car_data.get('total', 0),
        })
    
    # Добавляем итоговую строку
    totals = report_data.get('totals', {})
    rows.append({
        'Машина (номер)': 'ИТОГО',
        'Стоимость запчастей (сом)': totals.get('part_total', 0),
        'Стоимость работ (сом)': totals.get('job_total', 0),
        'Итого (сом)': totals.get('total', 0),
    })
    
    return rows


def prepare_fuel_consumption_for_export(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Подготовка данных отчета по топливу для экспорта.
    """
    rows = []
    
    for car_data in report_data.get('by_car', []):
        rows.append({
            'Машина (номер)': car_data.get('car__numplate', ''),
            'Марка': car_data.get('car__brand', ''),
            'Литры': car_data.get('total_liters', 0),
            'Стоимость (сом)': car_data.get('total_cost', 0),
            'Пробег (км)': car_data.get('total_mileage', 0),
            'Средний расход (л/100км)': car_data.get('avg_consumption', 0),
        })
    
    totals = report_data.get('totals', {})
    rows.append({
        'Машина (номер)': 'ИТОГО',
        'Марка': '',
        'Литры': totals.get('total_liters', 0),
        'Стоимость (сом)': totals.get('total_cost', 0),
        'Пробег (км)': totals.get('total_mileage', 0),
        'Средний расход (л/100км)': totals.get('avg_consumption', 0),
    })
    
    return rows


def prepare_insurance_inspection_for_export(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Подготовка данных отчета по страховкам/техосмотрам для экспорта.
    """
    rows = []
    
    status_map = {
        'active': 'Активна',
        'expiring_soon': 'Скоро истекает',
        'expired': 'Просрочена',
    }
    
    type_map = {
        'insurance': 'Страховка',
        'inspection': 'Техосмотр',
    }
    
    for item in report_data.get('items', []):
        rows.append({
            'Тип': type_map.get(item.get('type', ''), item.get('type', '')),
            'Машина (номер)': item.get('car__numplate', ''),
            'Номер': item.get('number', ''),
            'Дата начала': item.get('start_date', ''),
            'Дата окончания': item.get('end_date', ''),
            'Стоимость (сом)': item.get('cost', 0),
            'Статус': status_map.get(item.get('status', ''), item.get('status', '')),
        })

    return rows


def prepare_vehicle_utilization_for_export(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Подготовка данных отчета по утилизации транспорта для экспорта.
    """
    rows = []

    for car_data in report_data.get('data', []):
        rows.append({
            'Машина (номер)': car_data.get('car_numplate', ''),
            'Общий пробег (км)': car_data.get('total_mileage', 0),
            'Месяцев учёта': car_data.get('months', 0),
            'Средний пробег в месяц (км)': car_data.get('avg_monthly_mileage', 0),
        })

    return rows


def prepare_cost_analysis_for_export(report_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Подготовка данных отчета по анализу затрат для экспорта.
    """
    rows = []

    for car_data in report_data.get('data', []):
        rows.append({
            'Машина (номер)': car_data.get('car_numplate', ''),
            'Топливо (сом)': car_data.get('fuel_cost', 0),
            'ТО и ремонты (сом)': car_data.get('maintenance_cost', 0),
            'Страховка (сом)': car_data.get('insurance_cost', 0),
            'Техосмотр (сом)': car_data.get('inspection_cost', 0),
            'Итого (сом)': car_data.get('total_cost', 0),
        })

    # Добавляем итоговую строку
    summary = report_data.get('summary', {})
    rows.append({
        'Машина (номер)': 'ИТОГО',
        'Топливо (сом)': summary.get('total_fuel_cost', 0),
        'ТО и ремонты (сом)': summary.get('total_maintenance_cost', 0),
        'Страховка (сом)': summary.get('total_insurance_cost', 0),
        'Техосмотр (сом)': summary.get('total_inspection_cost', 0),
        'Итого (сом)': summary.get('grand_total', 0),
    })

    return rows


def get_export_preparator(report_type: str):
    """
    Возвращает функцию для подготовки данных отчета к экспорту.
    
    Args:
        report_type: Тип отчета
        
    Returns:
        Функция для подготовки данных или None
    """
    preparators = {
        'fuel_consumption': prepare_fuel_consumption_for_export,
        'maintenance_costs': prepare_maintenance_costs_for_export,
        'insurance_inspection': prepare_insurance_inspection_for_export,
        'vehicle_utilization': prepare_vehicle_utilization_for_export,
        'cost_analysis': prepare_cost_analysis_for_export,
    }
    
    return preparators.get(report_type)
